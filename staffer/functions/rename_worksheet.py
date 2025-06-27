"""Rename Excel worksheet function."""

import os
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def rename_worksheet(working_directory, filepath, old_name, new_name):
    """Rename worksheet in workbook.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to workbook file, relative to working directory
        old_name: Current name of the worksheet
        new_name: New name for the worksheet
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    # Check if workbook file exists
    if not os.path.exists(file_abs_path):
        return f'Error: Workbook "{filepath}" does not exist'
    
    try:
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if old sheet exists
        if old_name not in wb.sheetnames:
            return f'Error: Sheet "{old_name}" not found in workbook'
        
        # Check if new name already exists
        if new_name in wb.sheetnames:
            return f'Error: Sheet "{new_name}" already exists in workbook'
        
        # Rename sheet
        sheet = wb[old_name]
        sheet.title = new_name
        
        # Save workbook
        wb.save(file_abs_path)
        
        return f"Successfully renamed sheet from '{old_name}' to '{new_name}' in {filepath}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_rename_worksheet = types.FunctionDeclaration(
    name="rename_worksheet",
    description="Rename a worksheet in an Excel workbook",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the workbook file, relative to the working directory",
            ),
            "old_name": types.Schema(
                type=types.Type.STRING,
                description="Current name of the worksheet to rename",
            ),
            "new_name": types.Schema(
                type=types.Type.STRING,
                description="New name for the worksheet",
            ),
        },
        required=["filepath", "old_name", "new_name"],
    ),
)