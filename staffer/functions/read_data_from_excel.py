"""Read data from Excel worksheet function."""

import os
import json
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def read_data_from_excel(working_directory, filepath, sheet_name, start_cell="A1", end_cell=None):
    """Read data from Excel worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet to read from
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        
    Returns:
        String containing the data or error message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path, read_only=True)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # If no end_cell specified, try to find used range
        if end_cell is None:
            # Find the maximum row and column with data
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row == 1 and max_col == 1:
                # Check if A1 has data
                if ws['A1'].value is None:
                    return "No data found in worksheet"
                end_cell = "A1"
            else:
                # Convert column number to letter
                from openpyxl.utils import get_column_letter
                end_cell = f"{get_column_letter(max_col)}{max_row}"
        
        # Parse cell range
        try:
            cell_range = f"{start_cell}:{end_cell}" if end_cell != start_cell else start_cell
            cells = ws[cell_range]
        except Exception as e:
            return f"Error: Invalid cell range {start_cell}:{end_cell} - {str(e)}"
        
        # Extract data
        data = []
        if isinstance(cells, tuple):
            # Multiple rows
            for row in cells:
                if isinstance(row, tuple):
                    # Multiple columns in row
                    row_data = [cell.value for cell in row]
                else:
                    # Single column
                    row_data = [row.value]
                data.append(row_data)
        else:
            # Single cell
            data = [[cells.value]]
        
        wb.close()
        
        # Format result as JSON string
        result = {
            "sheet_name": sheet_name,
            "range": f"{start_cell}:{end_cell}",
            "data": data,
            "rows": len(data),
            "columns": len(data[0]) if data else 0
        }
        
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_read_data_from_excel = types.FunctionDeclaration(
    name="read_data_from_excel",
    description="Read data from Excel worksheet with optional cell range",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of worksheet to read from",
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Starting cell (default A1)",
            ),
            "end_cell": types.Schema(
                type=types.Type.STRING,
                description="Ending cell (optional, auto-expands if not provided)",
            ),
        },
        required=["filepath", "sheet_name"],
    ),
)