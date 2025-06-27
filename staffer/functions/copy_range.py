"""Copy range of cells in Excel worksheet function."""

import os
from copy import copy as copy_style
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def copy_range(working_directory, filepath, sheet_name, source_start, source_end, target_start, target_sheet=None):
    """Copy a range of cells to another location.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of source worksheet
        source_start: Top-left cell of source range (e.g., "A1")
        source_end: Bottom-right cell of source range (e.g., "B2")
        target_start: Top-left cell where to paste (e.g., "D5")
        target_sheet: Name of target worksheet (optional, defaults to source sheet)
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if source sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        source_ws = wb[sheet_name]
        
        # Determine target worksheet
        if target_sheet:
            if target_sheet not in wb.sheetnames:
                return f"Error: Target sheet '{target_sheet}' not found in workbook"
            target_ws = wb[target_sheet]
        else:
            target_ws = source_ws
        
        # Parse cell references
        try:
            from openpyxl.utils.cell import coordinate_from_string
            
            # Parse source range
            source_col_start, source_row_start = coordinate_from_string(source_start)
            source_col_end, source_row_end = coordinate_from_string(source_end)
            source_col_start_idx = column_index_from_string(source_col_start)
            source_col_end_idx = column_index_from_string(source_col_end)
            
            # Parse target start
            target_col, target_row = coordinate_from_string(target_start)
            target_col_idx = column_index_from_string(target_col)
            
        except Exception as e:
            return f"Error: Invalid cell reference - {str(e)}"
        
        # Calculate offsets
        row_offset = target_row - source_row_start
        col_offset = target_col_idx - source_col_start_idx
        
        # Copy cells
        cells_copied = 0
        for row in range(source_row_start, source_row_end + 1):
            for col_idx in range(source_col_start_idx, source_col_end_idx + 1):
                # Get source cell
                source_cell = source_ws.cell(row=row, column=col_idx)
                
                # Calculate target position
                target_row_pos = row + row_offset
                target_col_pos = col_idx + col_offset
                
                # Get target cell
                target_cell = target_ws.cell(row=target_row_pos, column=target_col_pos)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy style if cell has style
                if source_cell.has_style:
                    target_cell._style = copy_style(source_cell._style)
                
                cells_copied += 1
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        # Format result message
        if target_sheet and target_sheet != sheet_name:
            return f"Range copied successfully ({cells_copied} cells) from sheet '{sheet_name}' to sheet '{target_sheet}'"
        else:
            return f"Range copied successfully ({cells_copied} cells)"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_copy_range = types.FunctionDeclaration(
    name="copy_range",
    description="Copy a range of cells to another location",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of source worksheet",
            ),
            "source_start": types.Schema(
                type=types.Type.STRING,
                description="Top-left cell of source range (e.g., 'A1')",
            ),
            "source_end": types.Schema(
                type=types.Type.STRING,
                description="Bottom-right cell of source range (e.g., 'B2')",
            ),
            "target_start": types.Schema(
                type=types.Type.STRING,
                description="Top-left cell where to paste (e.g., 'D5')",
            ),
            "target_sheet": types.Schema(
                type=types.Type.STRING,
                description="Name of target worksheet (optional, defaults to source sheet)",
            ),
        },
        required=["filepath", "sheet_name", "source_start", "source_end", "target_start"],
    ),
)