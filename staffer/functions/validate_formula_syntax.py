"""Validate Excel formula syntax function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def validate_formula_syntax(working_directory, filepath, sheet_name, cell, formula):
    """Validate Excel formula syntax without applying it.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        cell: Cell reference (e.g., 'A1')
        formula: Formula to validate (with or without '=' prefix)
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File '{filepath}' does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        # Validate cell reference format
        if not _validate_cell_reference(cell):
            return f"Error: Invalid cell reference '{cell}'"
        
        # Validate formula syntax
        is_valid, message = _validate_formula(formula)
        if not is_valid:
            return f"Error: Invalid formula syntax: {message}"
        
        # Get worksheet and check current cell content
        sheet = wb[sheet_name]
        cell_obj = sheet[cell]
        current_value = cell_obj.value
        
        # Format response based on cell content
        if isinstance(current_value, str) and current_value.startswith('='):
            # Cell has a formula
            formula_normalized = formula if formula.startswith('=') else f'={formula}'
            if current_value == formula_normalized:
                return f"Formula is valid and matches cell {cell} content"
            else:
                return f"Formula is valid but doesn't match cell {cell} content (current: {current_value})"
        else:
            # Cell has no formula
            return f"Formula is valid but cell {cell} contains no formula (current: {current_value})"
        
    except Exception as e:
        return f"Error: {str(e)}"


def _validate_cell_reference(cell):
    """Validate cell reference format (e.g., A1, B2, etc.)."""
    import re
    pattern = r'^[A-Z]+[0-9]+$'
    return bool(re.match(pattern, cell))


def _validate_formula(formula):
    """Validate Excel formula syntax."""
    # Handle formula with or without '=' prefix
    if not formula.startswith('='):
        formula = f'={formula}'
    
    # Remove the '=' prefix for validation
    formula_body = formula[1:]
    
    # Check for balanced parentheses
    parens = 0
    for c in formula_body:
        if c == "(":
            parens += 1
        elif c == ")":
            parens -= 1
        if parens < 0:
            return False, "Unmatched closing parenthesis"
    
    if parens > 0:
        return False, "Unclosed parenthesis"
    
    # Check for basic syntax errors
    if formula_body.strip() == "":
        return False, "Empty formula"
    
    # Check for double operators
    import re
    if re.search(r'[+\-*/]{2,}', formula_body):
        return False, "Invalid operator sequence"
    
    # Basic function name validation
    func_pattern = r"([A-Z]+)\("
    funcs = re.findall(func_pattern, formula_body)
    unsafe_funcs = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}
    
    for func in funcs:
        if func in unsafe_funcs:
            return False, f"Unsafe function: {func}"
    
    return True, "Formula is valid"


# Schema for Google AI function declaration
schema_validate_formula_syntax = types.FunctionDeclaration(
    name="validate_formula_syntax",
    description="Validate Excel formula syntax without applying it to the worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet",
            ),
            "cell": types.Schema(
                type=types.Type.STRING,
                description="Cell reference (e.g., 'A1', 'B2')",
            ),
            "formula": types.Schema(
                type=types.Type.STRING,
                description="Excel formula to validate (with or without '=' prefix)",
            ),
        },
        required=["filepath", "sheet_name", "cell", "formula"],
    ),
)