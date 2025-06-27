"""Apply Excel formula to cell function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def apply_formula(working_directory, filepath, sheet_name, cell, formula):
    """Apply Excel formula to cell.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        cell: Cell reference (e.g., 'A1')
        formula: Formula to apply (with or without '=' prefix)
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File '{filepath}' does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        # Validate cell reference format
        if not _validate_cell_reference(cell):
            return f"Error: Invalid cell reference '{cell}'"
        
        # Validate formula syntax first
        is_valid, message = _validate_formula(formula)
        if not is_valid:
            return f"Error: Invalid formula syntax: {message}"
        
        # Ensure formula starts with '='
        formula_normalized = formula if formula.startswith('=') else f'={formula}'
        
        # Get worksheet and apply formula
        sheet = wb[sheet_name]
        cell_obj = sheet[cell]
        cell_obj.value = formula_normalized
        
        # Save workbook
        wb.save(file_abs_path)
        
        return f"Applied formula '{formula_normalized}' to cell {cell}"
        
    except Exception as e:
        return f"Error: {str(e)}"


def _validate_cell_reference(cell):
    """Validate cell reference format (e.g., A1, B2, etc.)."""
    import re
    pattern = r'^[A-Z]+[0-9]+$'
    return bool(re.match(pattern, cell))


def _validate_formula(formula):
    """Validate Excel formula syntax."""
    # Handle formula with or without '=' prefix
    if not formula.startswith('='):
        formula = f'={formula}'
    
    # Remove the '=' prefix for validation
    formula_body = formula[1:]
    
    # Check for balanced parentheses
    parens = 0
    for c in formula_body:
        if c == "(":
            parens += 1
        elif c == ")":
            parens -= 1
        if parens < 0:
            return False, "Unmatched closing parenthesis"
    
    if parens > 0:
        return False, "Unclosed parenthesis"
    
    # Check for basic syntax errors
    if formula_body.strip() == "":
        return False, "Empty formula"
    
    # Check for double operators
    import re
    if re.search(r'[+\-*/]{2,}', formula_body):
        return False, "Invalid operator sequence"
    
    # Basic function name validation
    func_pattern = r"([A-Z]+)\("
    funcs = re.findall(func_pattern, formula_body)
    unsafe_funcs = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}
    
    for func in funcs:
        if func in unsafe_funcs:
            return False, f"Unsafe function: {func}"
    
    return True, "Formula is valid"


# Schema for Google AI function declaration
schema_apply_formula = types.FunctionDeclaration(
    name="apply_formula",
    description="Apply Excel formula to a specific cell in a worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet",
            ),
            "cell": types.Schema(
                type=types.Type.STRING,
                description="Cell reference (e.g., 'A1', 'B2')",
            ),
            "formula": types.Schema(
                type=types.Type.STRING,
                description="Excel formula to apply (with or without '=' prefix)",
            ),
        },
        required=["filepath", "sheet_name", "cell", "formula"],
    ),
)