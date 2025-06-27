"""Get Excel workbook metadata function."""

import os
import json
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def get_workbook_metadata(working_directory, filepath, include_ranges=False):
    """Get metadata about workbook including sheets, ranges, etc.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to workbook file, relative to working directory
        include_ranges: Whether to include used ranges for each sheet
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    # Check if workbook file exists
    if not os.path.exists(file_abs_path):
        return f'Error: Workbook "{filepath}" does not exist'
    
    try:
        # Load workbook
        wb = load_workbook(file_abs_path, read_only=True)
        
        # Get file info
        path = Path(file_abs_path)
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                else:
                    ranges[sheet_name] = "No data"
            info["used_ranges"] = ranges
        
        wb.close()
        
        # Return formatted metadata
        return json.dumps(info, indent=2)
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_get_workbook_metadata = types.FunctionDeclaration(
    name="get_workbook_metadata",
    description="Get metadata about an Excel workbook including sheets, size, and optionally used ranges",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the workbook file, relative to the working directory",
            ),
            "include_ranges": types.Schema(
                type=types.Type.BOOLEAN,
                description="Whether to include used ranges for each sheet (default: false)",
            ),
        },
        required=["filepath"],
    ),
)