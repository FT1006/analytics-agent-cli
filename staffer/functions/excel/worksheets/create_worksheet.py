"""Create Excel worksheet function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_worksheet(working_directory, filepath, sheet_name):
    """Create new worksheet in existing Excel workbook.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to workbook file, relative to working directory
        sheet_name: Name for the new worksheet
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    # Check if workbook file exists
    if not os.path.exists(file_abs_path):
        return f'Error: Workbook "{filepath}" does not exist'
    
    try:
        # Load existing workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet name already exists
        if sheet_name in wb.sheetnames:
            return f'Error: Worksheet "{sheet_name}" already exists in workbook'
        
        # Create new worksheet
        wb.create_sheet(sheet_name)
        
        # Save workbook
        wb.save(file_abs_path)
        
        return f"Created worksheet '{sheet_name}' in {filepath}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_create_worksheet = types.FunctionDeclaration(
    name="create_worksheet",
    description="Create a new worksheet in an existing Excel workbook",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the workbook file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name for the new worksheet",
            ),
        },
        required=["filepath", "sheet_name"],
    ),
)