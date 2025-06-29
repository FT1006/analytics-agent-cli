"""Create Excel table function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_table(working_directory, filepath, sheet_name, table_name, data_range):
    """Create native Excel table from a specified range of data.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        table_name: Name for the table
        data_range: Cell range for the table (e.g., "A1:C10")
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Validate data range format
        if ":" not in data_range:
            wb.close()
            return f"Error: Invalid data range format '{data_range}'. Expected format like 'A1:C10'"
        
        # Check if table name already exists
        existing_tables = [table.name for table in ws._tables]
        if table_name in existing_tables:
            wb.close()
            return f"Error: Table '{table_name}' already exists in sheet '{sheet_name}'"
        
        # Create table
        table = Table(displayName=table_name, ref=data_range)
        
        # Add default table style
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Add table to worksheet
        ws.add_table(table)
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Table '{table_name}' created successfully in sheet '{sheet_name}' with range {data_range}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_create_table = types.FunctionDeclaration(
    name="create_table",
    description="Create a native Excel table from a specified range of data",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet",
            ),
            "table_name": types.Schema(
                type=types.Type.STRING,
                description="Name for the table",
            ),
            "data_range": types.Schema(
                type=types.Type.STRING,
                description="Cell range for the table (e.g., 'A1:C10')",
            ),
        },
        required=["filepath", "sheet_name", "table_name", "data_range"],
    ),
)