"""Create Excel chart function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_chart(working_directory, filepath, sheet_name, data_range, chart_type, target_cell, title=""):
    """Create chart in worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        data_range: Cell range for chart data (e.g., "A1:B10")
        chart_type: Type of chart ("column", "bar", "line", "pie", "area", "scatter")
        target_cell: Cell where chart will be placed (e.g., "D2")
        title: Chart title (optional)
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Validate data range format
        if ":" not in data_range:
            wb.close()
            return f"Error: Invalid data range format '{data_range}'. Expected format like 'A1:B10'"
        
        # Create chart based on type
        chart_type = chart_type.lower()
        if chart_type in ["column", "bar"]:
            chart = BarChart()
            if chart_type == "column":
                chart.type = "col"
            else:
                chart.type = "bar"
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "area":
            chart = AreaChart()
        elif chart_type == "scatter":
            chart = ScatterChart()
        else:
            wb.close()
            return f"Error: Unsupported chart type '{chart_type}'. Supported types: column, bar, line, pie, area, scatter"
        
        # Set chart title
        if title:
            chart.title = title
        
        # Parse data range
        start_cell, end_cell = data_range.split(":")
        start_col_letter = start_cell[0]
        start_row = int(start_cell[1:])
        end_col_letter = end_cell[0]
        end_row = int(end_cell[1:])
        
        start_col = ord(start_col_letter) - ord('A') + 1
        end_col = ord(end_col_letter) - ord('A') + 1
        
        # Create data reference
        data = Reference(ws, min_col=start_col, min_row=start_row, max_col=end_col, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        
        # For non-pie charts, set category labels from first column
        if chart_type != "pie" and start_row < end_row:  # Multiple rows, use first column as categories
            cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=end_row, max_col=start_col)
            chart.set_categories(cats)
        
        # Position chart at target cell
        chart.anchor = target_cell
        
        # Add chart to worksheet
        ws.add_chart(chart)
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Chart '{chart_type}' created successfully in sheet '{sheet_name}' at {target_cell}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_create_chart = types.FunctionDeclaration(
    name="create_chart_excel",
    description="Create and embed a chart in an Excel worksheet using cell ranges",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet",
            ),
            "data_range": types.Schema(
                type=types.Type.STRING,
                description="Cell range for chart data (e.g., 'A1:B10')",
            ),
            "chart_type": types.Schema(
                type=types.Type.STRING,
                description="Type of chart ('column', 'bar', 'line', 'pie', 'area', 'scatter')",
            ),
            "target_cell": types.Schema(
                type=types.Type.STRING,
                description="Cell where chart will be placed (e.g., 'D2')",
            ),
            "title": types.Schema(
                type=types.Type.STRING,
                description="Chart title (optional)",
            ),
        },
        required=["filepath", "sheet_name", "data_range", "chart_type", "target_cell"],
    ),
)