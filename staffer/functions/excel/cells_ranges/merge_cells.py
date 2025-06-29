"""Merge cells in Excel worksheet function."""

import os
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def merge_cells(working_directory, filepath, sheet_name, start_cell, end_cell):
    """Merge a range of cells in Excel worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet containing cells to merge
        start_cell: Top-left cell of range to merge (e.g., "A1")
        end_cell: Bottom-right cell of range to merge (e.g., "B2")
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Parse and validate cell references
        try:
            from openpyxl.utils.cell import coordinate_from_string
            # Validate start cell
            coordinate_from_string(start_cell)
            # Validate end cell
            coordinate_from_string(end_cell)
        except Exception as e:
            return f"Error: Invalid cell reference - {str(e)}"
        
        # Create range string
        range_string = f"{start_cell}:{end_cell}"
        
        # Check if range is already merged
        for merged_range in ws.merged_cells.ranges:
            if str(merged_range) == range_string:
                return f"Error: Range {range_string} is already merged"
        
        # Merge the cells
        ws.merge_cells(range_string)
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Range {range_string} merged in sheet '{sheet_name}'"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_merge_cells = types.FunctionDeclaration(
    name="merge_cells",
    description="Merge a range of cells in Excel worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of worksheet containing cells to merge",
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Top-left cell of range to merge (e.g., 'A1')",
            ),
            "end_cell": types.Schema(
                type=types.Type.STRING,
                description="Bottom-right cell of range to merge (e.g., 'B2')",
            ),
        },
        required=["filepath", "sheet_name", "start_cell", "end_cell"],
    ),
)