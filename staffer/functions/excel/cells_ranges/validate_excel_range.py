"""Validate Excel range function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def validate_excel_range(working_directory, filepath, sheet_name, start_cell, end_cell=None):
    """Validate if a range exists and is properly formatted.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet to validate
        start_cell: Starting cell of range
        end_cell: Ending cell of range (optional)
        
    Returns:
        String validation result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path, read_only=True)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Validate start cell format
        try:
            from openpyxl.utils.cell import coordinate_from_string
            from openpyxl.utils import column_index_from_string
            start_col_letter, start_row = coordinate_from_string(start_cell)
            start_col_idx = column_index_from_string(start_col_letter)
        except Exception as e:
            return f"Error: Invalid start cell format '{start_cell}' - {str(e)}"
        
        # Validate end cell format if provided
        end_col_idx = start_col_idx
        end_row = start_row
        
        if end_cell:
            try:
                end_col_letter, end_row = coordinate_from_string(end_cell)
                end_col_idx = column_index_from_string(end_col_letter)
            except Exception as e:
                return f"Error: Invalid end cell format '{end_cell}' - {str(e)}"
        
        # Check if range is logical (start <= end)
        if end_row < start_row or end_col_idx < start_col_idx:
            return f"Error: Invalid range - end cell ({end_cell}) must be after start cell ({start_cell})"
        
        # Check if range is within worksheet bounds
        max_row = ws.max_row
        max_col = ws.max_column
        
        # For empty worksheets, allow reasonable limits
        if max_row == 1 and max_col == 1 and ws['A1'].value is None:
            max_row = 1048576  # Excel row limit
            max_col = 16384    # Excel column limit
        
        if start_row > max_row or end_row > max_row:
            return f"Error: Row {max(start_row, end_row)} exceeds worksheet bounds (max row: {max_row})"
        
        if start_col_idx > max_col or end_col_idx > max_col:
            from openpyxl.utils import get_column_letter
            max_col_letter = get_column_letter(max_col)
            invalid_col_letter = get_column_letter(max(start_col_idx, end_col_idx))
            return f"Error: Column {invalid_col_letter} exceeds worksheet bounds (max column: {max_col_letter})"
        
        # Calculate range statistics
        range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
        num_rows = end_row - start_row + 1
        num_cols = end_col_idx - start_col_idx + 1
        total_cells = num_rows * num_cols
        
        wb.close()
        
        return f"Range '{range_str}' in sheet '{sheet_name}' is valid - {num_rows} rows x {num_cols} columns ({total_cells} cells)"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_validate_excel_range = types.FunctionDeclaration(
    name="validate_excel_range",
    description="Validate if a range exists and is properly formatted",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of worksheet to validate",
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Starting cell of range",
            ),
            "end_cell": types.Schema(
                type=types.Type.STRING,
                description="Ending cell of range (optional)",
            ),
        },
        required=["filepath", "sheet_name", "start_cell"],
    ),
)