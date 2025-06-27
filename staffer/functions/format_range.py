"""Format Excel range function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def format_range(working_directory, filepath, sheet_name, start_cell, end_cell=None, 
                bold=False, italic=False, underline=False, font_size=None, 
                font_color=None, bg_color=None, border_style=None, border_color=None, 
                number_format=None, alignment=None, wrap_text=False, merge_cells=False):
    """Apply formatting to a range of cells.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        start_cell: Starting cell reference (e.g., "A1")
        end_cell: Ending cell reference (optional, defaults to single cell)
        bold: Apply bold formatting
        italic: Apply italic formatting
        underline: Apply underline formatting
        font_size: Font size in points
        font_color: Font color (hex format like "FF0000" for red)
        bg_color: Background color (hex format like "FFFF00" for yellow)
        border_style: Border style ("thin", "thick", "medium", etc.)
        border_color: Border color (hex format)
        number_format: Number format string (e.g., "0.00", "0%")
        alignment: Text alignment ("left", "center", "right")
        wrap_text: Enable text wrapping
        merge_cells: Merge the specified range
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Handle single cell vs range
        if end_cell is None:
            cell_range = ws[start_cell]
            range_str = start_cell
        else:
            cell_range = ws[f"{start_cell}:{end_cell}"]
            range_str = f"{start_cell}:{end_cell}"
        
        # Create font style
        font_kwargs = {}
        if bold:
            font_kwargs['bold'] = True
        if italic:
            font_kwargs['italic'] = True
        if underline:
            font_kwargs['underline'] = 'single'
        if font_size:
            font_kwargs['size'] = font_size
        if font_color:
            font_kwargs['color'] = font_color
        
        font = Font(**font_kwargs) if font_kwargs else None
        
        # Create fill (background color)
        fill = None
        if bg_color:
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Create border
        border = None
        if border_style:
            side_style = Side(style=border_style, color=border_color or "000000")
            border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
        
        # Create alignment
        align = None
        if alignment or wrap_text:
            align_kwargs = {}
            if alignment:
                align_kwargs['horizontal'] = alignment
            if wrap_text:
                align_kwargs['wrap_text'] = True
            align = Alignment(**align_kwargs)
        
        # Apply formatting to cells
        if isinstance(cell_range, tuple):
            # Range of cells
            for row in cell_range:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if border:
                        cell.border = border
                    if align:
                        cell.alignment = align
                    if number_format:
                        cell.number_format = number_format
        else:
            # Single cell
            if font:
                cell_range.font = font
            if fill:
                cell_range.fill = fill
            if border:
                cell_range.border = border
            if align:
                cell_range.alignment = align
            if number_format:
                cell_range.number_format = number_format
        
        # Handle cell merging
        if merge_cells and end_cell:
            ws.merge_cells(f"{start_cell}:{end_cell}")
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Range {range_str} formatted successfully in sheet '{sheet_name}'"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_format_range = types.FunctionDeclaration(
    name="format_range",
    description="Apply formatting to a range of cells in an Excel worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet",
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Starting cell reference (e.g., 'A1')",
            ),
            "end_cell": types.Schema(
                type=types.Type.STRING,
                description="Ending cell reference (optional, defaults to single cell)",
            ),
            "bold": types.Schema(
                type=types.Type.BOOLEAN,
                description="Apply bold formatting",
            ),
            "italic": types.Schema(
                type=types.Type.BOOLEAN,
                description="Apply italic formatting",
            ),
            "underline": types.Schema(
                type=types.Type.BOOLEAN,
                description="Apply underline formatting",
            ),
            "font_size": types.Schema(
                type=types.Type.INTEGER,
                description="Font size in points",
            ),
            "font_color": types.Schema(
                type=types.Type.STRING,
                description="Font color in hex format (e.g., 'FF0000' for red)",
            ),
            "bg_color": types.Schema(
                type=types.Type.STRING,
                description="Background color in hex format (e.g., 'FFFF00' for yellow)",
            ),
            "border_style": types.Schema(
                type=types.Type.STRING,
                description="Border style ('thin', 'thick', 'medium', etc.)",
            ),
            "border_color": types.Schema(
                type=types.Type.STRING,
                description="Border color in hex format",
            ),
            "number_format": types.Schema(
                type=types.Type.STRING,
                description="Number format string (e.g., '0.00', '0%')",
            ),
            "alignment": types.Schema(
                type=types.Type.STRING,
                description="Text alignment ('left', 'center', 'right')",
            ),
            "wrap_text": types.Schema(
                type=types.Type.BOOLEAN,
                description="Enable text wrapping",
            ),
            "merge_cells": types.Schema(
                type=types.Type.BOOLEAN,
                description="Merge the specified range",
            ),
        },
        required=["filepath", "sheet_name", "start_cell"],
    ),
)