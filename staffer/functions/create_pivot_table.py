"""Create Excel pivot table function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_pivot_table(working_directory, filepath, sheet_name, data_range, pivot_sheet_name, 
                      row_fields, column_fields, value_fields):
    """Create pivot table in worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of source worksheet containing data
        data_range: Cell range for source data (e.g., "A1:D100")
        pivot_sheet_name: Name of new sheet for pivot table
        row_fields: List of field names for row labels
        column_fields: List of field names for column labels
        value_fields: List of field names for values
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if source sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Source sheet '{sheet_name}' not found in workbook"
        
        source_ws = wb[sheet_name]
        
        # Validate data range format
        if ":" not in data_range:
            wb.close()
            return f"Error: Invalid data range format '{data_range}'. Expected format like 'A1:D100'"
        
        # Get header row to validate field names
        start_cell, end_cell = data_range.split(":")
        start_col_idx = ord(start_cell[0]) - ord('A') + 1
        start_row_idx = int(start_cell[1:])
        end_col_idx = ord(end_cell[0]) - ord('A') + 1
        
        # Extract headers from first row of data range
        headers = []
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell_value = source_ws.cell(row=start_row_idx, column=col_idx).value
            if cell_value:
                headers.append(str(cell_value))
        
        # Validate that all field names exist in headers
        all_fields = row_fields + column_fields + value_fields
        invalid_fields = [field for field in all_fields if field not in headers]
        if invalid_fields:
            wb.close()
            return f"Error: Field(s) {invalid_fields} not found in data headers {headers}"
        
        # Create or get pivot sheet
        if pivot_sheet_name in wb.sheetnames:
            pivot_ws = wb[pivot_sheet_name]
        else:
            pivot_ws = wb.create_sheet(title=pivot_sheet_name)
        
        # Clear existing content
        pivot_ws.delete_rows(1, pivot_ws.max_row)
        
        # Create summary pivot table representation
        # Note: Full pivot table functionality is limited in openpyxl
        pivot_ws.cell(row=1, column=1, value="Pivot Table Summary")
        pivot_ws.cell(row=2, column=1, value=f"Source: {sheet_name}!{data_range}")
        pivot_ws.cell(row=3, column=1, value=f"Row Fields: {', '.join(row_fields)}")
        pivot_ws.cell(row=4, column=1, value=f"Column Fields: {', '.join(column_fields)}")
        pivot_ws.cell(row=5, column=1, value=f"Value Fields: {', '.join(value_fields)}")
        pivot_ws.cell(row=6, column=1, value="Note: Full pivot table creation requires Excel. This is a summary view.")
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Pivot table summary created in sheet '{pivot_sheet_name}' with rows: {row_fields}, columns: {column_fields}, values: {value_fields}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_create_pivot_table = types.FunctionDeclaration(
    name="create_pivot_table",
    description="Create a pivot table in an Excel worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of source worksheet containing data",
            ),
            "data_range": types.Schema(
                type=types.Type.STRING,
                description="Cell range for source data (e.g., 'A1:D100')",
            ),
            "pivot_sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of new sheet for pivot table",
            ),
            "row_fields": types.Schema(
                type=types.Type.ARRAY,
                items=types.Schema(type=types.Type.STRING),
                description="List of field names for row labels",
            ),
            "column_fields": types.Schema(
                type=types.Type.ARRAY,
                items=types.Schema(type=types.Type.STRING),
                description="List of field names for column labels",
            ),
            "value_fields": types.Schema(
                type=types.Type.ARRAY,
                items=types.Schema(type=types.Type.STRING),
                description="List of field names for values",
            ),
        },
        required=["filepath", "sheet_name", "data_range", "pivot_sheet_name", "row_fields", "column_fields", "value_fields"],
    ),
)