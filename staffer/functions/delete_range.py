"""Delete range of cells in Excel worksheet function."""

import os
from google.genai import types

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def delete_range(working_directory, filepath, sheet_name, start_cell, end_cell, shift_direction="up"):
    """Delete a range of cells and shift remaining cells.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet containing cells to delete
        start_cell: Top-left cell of range to delete (e.g., "A1")
        end_cell: Bottom-right cell of range to delete (e.g., "B2")
        shift_direction: Direction to shift cells after deletion ("up" or "left")
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Validate shift direction
        if shift_direction not in ["up", "left"]:
            return f"Error: Invalid shift direction: {shift_direction}. Must be 'up' or 'left'"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Parse cell references
        try:
            from openpyxl.utils.cell import coordinate_from_string
            
            # Parse range
            start_col, start_row = coordinate_from_string(start_cell)
            end_col, end_row = coordinate_from_string(end_cell)
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)
            
        except Exception as e:
            return f"Error: Invalid cell reference - {str(e)}"
        
        # Validate range within worksheet bounds
        if end_row > ws.max_row:
            return f"Error: End row {end_row} out of bounds (max row: {ws.max_row})"
        if end_col_idx > ws.max_column:
            return f"Error: End column {end_col} out of bounds (max column: {get_column_letter(ws.max_column)})"
        
        # Format range string for display
        range_string = f"{start_cell}:{end_cell}"
        
        # Clear cell contents and formatting first
        from openpyxl.styles import Font, Border, PatternFill
        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = None
                # Clear formatting - use default objects instead of None
                cell.font = Font()
                cell.border = Border()
                cell.fill = PatternFill()
                cell.number_format = "General"
                cell.alignment = None
        
        # Perform deletion with shift
        if shift_direction == "up":
            # Delete rows
            rows_to_delete = end_row - start_row + 1
            ws.delete_rows(start_row, rows_to_delete)
        else:  # shift_direction == "left"
            # Delete columns
            cols_to_delete = end_col_idx - start_col_idx + 1
            ws.delete_cols(start_col_idx, cols_to_delete)
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Range {range_string} deleted successfully"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_delete_range = types.FunctionDeclaration(
    name="delete_range",
    description="Delete a range of cells and shift remaining cells",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of worksheet containing cells to delete",
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Top-left cell of range to delete (e.g., 'A1')",
            ),
            "end_cell": types.Schema(
                type=types.Type.STRING,
                description="Bottom-right cell of range to delete (e.g., 'B2')",
            ),
            "shift_direction": types.Schema(
                type=types.Type.STRING,
                description="Direction to shift cells after deletion ('up' or 'left', default 'up')",
            ),
        },
        required=["filepath", "sheet_name", "start_cell", "end_cell"],
    ),
)