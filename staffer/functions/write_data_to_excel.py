"""Write data to Excel worksheet function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_data_to_excel(working_directory, filepath, sheet_name, data, start_cell="A1"):
    """Write data to Excel worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet to write to
        data: List of lists containing data to write (rows)
        start_cell: Cell to start writing to (default A1)
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Validate data format
        if not isinstance(data, list):
            return "Error: Data must be a list of rows"
        
        if not data:
            return "Error: Data cannot be empty"
        
        for i, row in enumerate(data):
            if not isinstance(row, list):
                return f"Error: Row {i} must be a list of values"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        ws = wb[sheet_name]
        
        # Parse starting cell
        try:
            from openpyxl.utils.cell import coordinate_from_string
            from openpyxl.utils import column_index_from_string
            col_letter, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)
        except Exception as e:
            return f"Error: Invalid start cell {start_cell} - {str(e)}"
        
        # Write data
        rows_written = 0
        cells_written = 0
        
        for row_idx, row_data in enumerate(data):
            current_row = start_row + row_idx
            for col_idx, cell_value in enumerate(row_data):
                current_col = start_col + col_idx
                ws.cell(row=current_row, column=current_col, value=cell_value)
                cells_written += 1
            rows_written += 1
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Successfully wrote {rows_written} rows ({cells_written} cells) to sheet '{sheet_name}' starting at {start_cell}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_write_data_to_excel = types.FunctionDeclaration(
    name="write_data_to_excel",
    description="Write data to Excel worksheet",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of worksheet to write to",
            ),
            "data": types.Schema(
                type=types.Type.ARRAY,
                description="List of lists containing data to write (rows)",
                items=types.Schema(
                    type=types.Type.ARRAY,
                    items=types.Schema(type=types.Type.STRING),
                ),
            ),
            "start_cell": types.Schema(
                type=types.Type.STRING,
                description="Cell to start writing to (default A1)",
            ),
        },
        required=["filepath", "sheet_name", "data"],
    ),
)