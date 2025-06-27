"""Get data validation info from Excel worksheet function."""

import os
import json
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def get_data_validation_info(working_directory, filepath, sheet_name):
    """Get all data validation rules in a worksheet.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        sheet_name: Name of worksheet
        
    Returns:
        String result message with validation info
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File '{filepath}' does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path, read_only=False)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook"
        
        # Get worksheet
        ws = wb[sheet_name]
        
        # Get all validation rules
        validations = _get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
        
        # Format result as JSON string
        result = {
            "sheet_name": sheet_name,
            "validation_rules": validations
        }
        
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        return f"Error: {str(e)}"


def _get_all_validation_ranges(worksheet):
    """Get all data validation ranges in a worksheet."""
    validations = []
    
    try:
        for dv in worksheet.data_validations.dataValidation:
            validation_info = {
                "ranges": str(dv.sqref),
                "validation_type": dv.type,
                "allow_blank": dv.allowBlank,
            }
            
            # Add operator for validation types that use it
            if dv.operator:
                validation_info["operator"] = dv.operator
            
            # Add optional fields if they exist
            if dv.prompt:
                validation_info["prompt"] = dv.prompt
            if dv.promptTitle:
                validation_info["prompt_title"] = dv.promptTitle
            if dv.error:
                validation_info["error_message"] = dv.error
            if dv.errorTitle:
                validation_info["error_title"] = dv.errorTitle
            
            # For list type validations (dropdown lists), extract allowed values
            if dv.type == "list" and dv.formula1:
                validation_info["allowed_values"] = _extract_list_values(dv.formula1, worksheet)
            
            # For other validation types, include the formulas
            elif dv.formula1:
                validation_info["formula1"] = dv.formula1
                if dv.formula2:
                    validation_info["formula2"] = dv.formula2
            
            validations.append(validation_info)
            
    except Exception as e:
        # Log warning but don't fail completely
        pass
    
    return validations


def _extract_list_values(formula, worksheet):
    """Extract allowed values from a list validation formula."""
    try:
        # Remove quotes if present
        formula = formula.strip('"')
        
        # Handle comma-separated list
        if ',' in formula:
            # Split by comma and clean up each value
            values = [val.strip().strip('"') for val in formula.split(',')]
            return [val for val in values if val]  # Remove empty values
        
        # Handle range reference (e.g., "$A$1:$A$5" or "Sheet1!$A$1:$A$5")
        elif (':' in formula or formula.startswith('$')) and worksheet:
            try:
                # Remove potential leading '=' if it's a formula like '=Sheet1!$A$1:$A$5'
                range_ref = formula
                if formula.startswith('='):
                    range_ref = formula[1:]
                
                actual_values = []
                # worksheet[range_ref] can resolve ranges like "A1:A5"
                range_cells = worksheet[range_ref]
                
                # Handle single cell or range
                if hasattr(range_cells, 'value'):  # Single cell
                    if range_cells.value is not None:
                        actual_values.append(str(range_cells.value))
                else:  # Range of cells
                    for row_of_cells in range_cells:
                        # Handle case where row_of_cells might be a single cell
                        if hasattr(row_of_cells, 'value'):
                            if row_of_cells.value is not None:
                                actual_values.append(str(row_of_cells.value))
                        else:
                            for cell in row_of_cells:
                                if cell.value is not None:
                                    actual_values.append(str(cell.value))
                
                if actual_values:
                    return actual_values
                return [f"Range: {formula} (empty or unresolvable)"]
                
            except Exception:
                return [f"Range: {formula} (resolution error)"]
        
        # Handle range reference when worksheet not available
        elif ':' in formula or formula.startswith('$'):
            return [f"Range: {formula}"]
        
        # Single value
        else:
            return [formula.strip('"')]
    
    except Exception:
        return [formula]  # Return original formula if parsing fails


# Schema for Google AI function declaration
schema_get_data_validation_info = types.FunctionDeclaration(
    name="get_data_validation_info",
    description="Get all data validation rules in a worksheet. Returns JSON with validation rule details including types, ranges, and constraints.",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the Excel file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet to analyze for validation rules",
            ),
        },
        required=["filepath", "sheet_name"],
    ),
)