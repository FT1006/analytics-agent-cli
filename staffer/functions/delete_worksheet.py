"""Delete Excel worksheet function."""

import os
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def delete_worksheet(working_directory, filepath, sheet_name):
    """Delete worksheet from workbook.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to workbook file, relative to working directory
        sheet_name: Name of the worksheet to delete
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    # Check if workbook file exists
    if not os.path.exists(file_abs_path):
        return f'Error: Workbook "{filepath}" does not exist'
    
    try:
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f'Error: Sheet "{sheet_name}" not found in workbook'
        
        # Check if it's the only sheet (can't delete the last sheet)
        if len(wb.sheetnames) == 1:
            return f'Error: Cannot delete "{sheet_name}" as it is the only sheet in the workbook'
        
        # Delete sheet
        del wb[sheet_name]
        
        # Save workbook
        wb.save(file_abs_path)
        
        return f"Successfully deleted sheet '{sheet_name}' from {filepath}"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_delete_worksheet = types.FunctionDeclaration(
    name="delete_worksheet",
    description="Delete a worksheet from an Excel workbook",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to the workbook file, relative to the working directory",
            ),
            "sheet_name": types.Schema(
                type=types.Type.STRING,
                description="Name of the worksheet to delete",
            ),
        },
        required=["filepath", "sheet_name"],
    ),
)