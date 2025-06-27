"""Copy worksheet function."""

import os
from pathlib import Path
from google.genai import types

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def copy_worksheet(working_directory, filepath, source_sheet, target_sheet):
    """Copy worksheet within workbook.
    
    Args:
        working_directory: The permitted working directory
        filepath: Path to Excel file, relative to working directory
        source_sheet: Name of source worksheet to copy
        target_sheet: Name of target worksheet to create
        
    Returns:
        String result message
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl library not available. Please install with: pip install openpyxl"
    
    # Security: Validate file path is within working directory
    working_dir_abs = os.path.abspath(working_directory)
    file_abs_path = os.path.abspath(os.path.join(working_dir_abs, filepath))
    
    if not file_abs_path.startswith(working_dir_abs):
        return f'Error: Cannot access "{filepath}" as it is outside the permitted working directory'
    
    try:
        # Check if file exists
        if not os.path.exists(file_abs_path):
            return f"Error: File {filepath} does not exist"
        
        # Load workbook
        wb = load_workbook(file_abs_path)
        
        # Check if source sheet exists
        if source_sheet not in wb.sheetnames:
            return f"Error: Source sheet '{source_sheet}' not found in workbook"
        
        # Check if target sheet already exists
        if target_sheet in wb.sheetnames:
            return f"Error: Target sheet '{target_sheet}' already exists in workbook"
        
        # Get source worksheet
        source_ws = wb[source_sheet]
        
        # Copy worksheet
        target_ws = wb.copy_worksheet(source_ws)
        target_ws.title = target_sheet
        
        # Save workbook
        wb.save(file_abs_path)
        wb.close()
        
        return f"Successfully copied sheet '{source_sheet}' to '{target_sheet}'"
        
    except Exception as e:
        return f"Error: {str(e)}"


# Schema for Google AI function declaration
schema_copy_worksheet = types.FunctionDeclaration(
    name="copy_worksheet",
    description="Copy worksheet within workbook",
    parameters=types.Schema(
        type=types.Type.OBJECT,
        properties={
            "filepath": types.Schema(
                type=types.Type.STRING,
                description="Path to Excel file, relative to the working directory",
            ),
            "source_sheet": types.Schema(
                type=types.Type.STRING,
                description="Name of source worksheet to copy",
            ),
            "target_sheet": types.Schema(
                type=types.Type.STRING,
                description="Name of target worksheet to create",
            ),
        },
        required=["filepath", "source_sheet", "target_sheet"],
    ),
)